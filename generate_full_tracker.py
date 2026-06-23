#!/usr/bin/env python3
"""
Generate the full G&A Benefits Platform Migration Effort Tracker.
Covers all 6 initiatives from the master Benefits PRD.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_PATH = "/Users/abhijeettiwari/Documents/ga-CAP-demo-dev/CAP_Migration_Effort_Tracker.xlsx"

# ── Color Palette (same as original) ──────────────────────────────────────────
DARK_HEADER_BG = "13212C"       # Dark navy for header rows
SECTION_BG     = "F5F7FA"       # Light gray for section headers
BORDER_COLOR   = "E4E8ED"       # Light border
TITLE_COLOR    = "C60C30"       # Red for titles / section text
SUBTITLE_COLOR = "64707A"       # Gray for subtitles
WHITE          = "FFFFFF"
HIGHLIGHT_GREEN = "E8F5E9"      # Light green for totals / highlights
HIGHLIGHT_BLUE  = "E3F2FD"      # Light blue for phase highlight
HIGHLIGHT_AMBER = "FFF8E1"      # Light amber for risk highlight
INIT_B_BG      = "F3E5F5"       # Light purple for Init B section
INIT_C_BG      = "E8EAF6"       # Light indigo for Init C section
INIT_D_BG      = "FFF3E0"       # Light orange for Init D section
INIT_E_BG      = "E0F2F1"       # Light teal for Init E section
INIT_F_BG      = "FCE4EC"       # Light pink for Init F section

# ── Reusable Styles ──────────────────────────────────────────────────────────
thin_border = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="thin", color=BORDER_COLOR),
)

title_font     = Font(name="Calibri", size=14, bold=True, color=TITLE_COLOR)
subtitle_font  = Font(name="Calibri", size=9, color=SUBTITLE_COLOR)
header_font    = Font(name="Calibri", size=11, bold=True, color=WHITE)
header_fill    = PatternFill(start_color=DARK_HEADER_BG, end_color=DARK_HEADER_BG, fill_type="solid")
header_align   = Alignment(horizontal="center", vertical="center", wrap_text=True)
section_font   = Font(name="Calibri", size=11, bold=True, color=TITLE_COLOR)
section_fill   = PatternFill(start_color=SECTION_BG, end_color=SECTION_BG, fill_type="solid")
data_font      = Font(name="Calibri", size=10)
data_align     = Alignment(vertical="top", wrap_text=True)
bold_font      = Font(name="Calibri", size=10, bold=True)
bold_font_11   = Font(name="Calibri", size=11, bold=True)
total_fill     = PatternFill(start_color=HIGHLIGHT_GREEN, end_color=HIGHLIGHT_GREEN, fill_type="solid")
note_font      = Font(name="Calibri", size=9, italic=True, color=SUBTITLE_COLOR)


def apply_header_row(ws, row, values):
    """Apply dark-header styling to a row."""
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def apply_data_row(ws, row, values, font=None, fill=None, align=None):
    """Apply data styling to a row."""
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = font or data_font
        cell.alignment = align or data_align
        cell.border = thin_border
        if fill:
            cell.fill = fill


def apply_section_row(ws, row, label, num_cols):
    """Apply section-header styling spanning all columns."""
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = section_font
    cell.fill = section_fill
    cell.border = thin_border
    for c in range(2, num_cols + 1):
        cl = ws.cell(row=row, column=c)
        cl.fill = section_fill
        cl.border = thin_border
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# =============================================================================
# SHEET 1: Migration Tracker
# =============================================================================
def build_migration_tracker(wb):
    ws = wb.active
    ws.title = "Migration Tracker"
    NUM_COLS = 11

    set_col_widths(ws, [5, 14, 28, 50, 20, 14, 14, 22, 10, 14, 35])

    # Title rows
    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NUM_COLS)
    ws.cell(r, 1, "G&A Benefits Platform — Full Migration Effort Tracker").font = title_font

    r = 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NUM_COLS)
    ws.cell(r, 1, "Ref: Master Benefits PRD — 6 Initiatives (A-F) · 5 Detailed PRDs (Foundation, New Business, Renewal, Copilot, Blueprint) · 22-Week Phased Delivery").font = subtitle_font

    r = 3  # blank row

    # Header row
    r = 4
    headers = ["#", "Initiative", "Feature / Work Item", "Description", "Priority",
               "Status", "Est. Days", "Dependencies", "Phase", "Owner", "Notes"]
    apply_header_row(ws, r, headers)

    # ─── Initiative A data (all 52 items from original) ───
    init_a_data = [
        # FOUNDATION section
        ("FOUNDATION — Admin Console (F1-F10)", None),
        (1, "A", "Admin Console Shell", "Top-level admin area with left-nav, role gating, tab navigation for all foundation modules", "P0 — Critical Path", "POC Done", 1, "Auth/RBAC", "Phase 1", "FE", "POC complete; prod needs SSO wiring only"),
        (2, "A", "Connections Hub — PrismHR", "PrismHR connector: auth, scheduled + on-demand sync, delta detection, field mapping, sync log", "P0 — Critical Path", "Partial", 8, "Prism API access", "Phase 1", "BE", "POC simulated; prod needs REST client + retry logic"),
        (3, "A", "Connections Hub — ClientSpace", "ClientSpace connector: case creation, approval routing, benefits page export", "P1 — High", "Partial", 4, "CS API access", "Phase 1", "BE", "Webhook integration + case object mapping"),
        (4, "A", "Connections Hub — DocuSign", "DocuSign connector: envelope creation, template mapping, signer routing, status callbacks", "P1 — High", "Partial", 3, "DocuSign API key", "Phase 1", "BE", "OAuth2 + envelope API — well-documented API"),
        (5, "A", "Connections Hub — WorkSight", "WorkSight/NextGen: enrollment portal data push, eligibility config", "P2 — Medium", "Design Only", 4, "WS API/import format", "Phase 1", "BE", "Needs SME call to confirm API mechanism"),
        (6, "A", "Ongoing Sync — Stream 1 (Master Data)", "Continuous sync of carrier rates, master plan catalog, plan-code crosswalks with versioning", "P0 — Critical Path", "Partial", 6, "F2 Prism connector", "Phase 1", "BE", "Diff-review, version history storage"),
        (7, "A", "Ongoing Sync — Stream 2 (Client Records)", "Client-master mirror + per-client on-demand refresh for renewal baseline", "P0 — Critical Path", "Partial", 5, "F2 Prism connector", "Phase 1", "BE", "Change-capture log, timestamps"),
        (8, "A", "Master Plan Library", "Browsable plan catalog synced from Prism: search, filter, rate tables by tier/year, rate-change diff", "P0 — Critical Path", "Partial", 5, "F2 Stream 1", "Phase 1", "FE+BE", "POC has browse + metrics; swap to real data"),
        (9, "A", "Pricing Stack", "Buckets per carrier, admin-factor tables, commission schedules, risk-factor formula — versioned", "P0 — Critical Path", "Partial", 4, "None", "Phase 1", "FE+BE", "POC has calculator; add version mgmt + approver"),
        (10, "A", "Parameters & Fees", "ACA threshold (year-versioned), deduction options, fee schedule", "P1 — High", "Partial", 2, "None", "Phase 1", "FE+BE", "Straightforward CRUD + effective-dating"),
        (11, "A", "Vocabularies", "Managed dropdown lists: class types, waiting periods, entity types, etc.", "P1 — High", "Partial", 2, "None", "Phase 1", "FE+BE", "POC has class types; replicate for all lists"),
        (12, "A", "Document Templates", "Template editor with merge-field mapping: ER Confirmation, EE SRA (EN+ES), Booklet, Prism payload", "P1 — High", "Partial", 7, "Template engine", "Phase 1", "FE+BE", "Real template engine (WeasyPrint/Puppeteer)"),
        (13, "A", "Validation Rules", "Typed rule library: completeness, source-data, cross-field, audit gates — real-time evaluation", "P0 — Critical Path", "Partial", 5, "F3, F4", "Phase 1", "BE", "POC has dynamic checks; productionize rule engine"),
        (14, "A", "Roles & Access", "Field-level perms (blue/yellow/green/red), role matrix, PHI/PII classification", "P1 — High", "Partial", 4, "Auth provider", "Phase 1", "FE+BE", "SSO claim mapping + field-level enforcement"),
        (15, "A", "Underwriting Intake", "Document upload + extraction config, field map, completeness check", "P1 — High", "Partial", 5, "AI extraction svc", "Phase 1", "FE+BE+AI", "Real extraction pipeline"),

        # NEW BUSINESS section
        ("NEW BUSINESS — CAP Build-out (NB1-NB9)", None),
        (16, "A", "Intake & Document Upload", "Seed fields + document checklist + drag-drop multi-file upload with type detection", "P0 — Critical Path", "POC Done", 1, "None", "Phase 1", "FE", "Done; add S3 storage + virus scan"),
        (17, "A", "AI Extraction & Completeness Gate", "Extraction -> structured fields with confidence -> completeness gate before assembly", "P0 — Critical Path", "Partial", 8, "AI extraction svc", "Phase 1", "BE+AI", "POC static; prod needs extraction pipeline"),
        (18, "A", "Pre-filled CAP Assembly", "Fuse 3 sources with provenance, confirm buttons, Prism sync refresh icons", "P0 — Critical Path", "POC Done", 2, "F3, F4, F6, NB2", "Phase 1", "FE", "POC fully interactive; add BE persistence"),
        (19, "A", "Plan Design & Rates", "Rate formula, contribution strategy, per-plan tier tables, ACA check", "P0 — Critical Path", "POC Done", 2, "F3, F4, F5", "Phase 1", "FE+BE", "POC calculates live; add BE rate engine"),
        (20, "A", "Readiness & Flagging", "Dynamic validation, Fix/Dismiss, progress bar, handoff gate", "P0 — Critical Path", "POC Done", 3, "F8", "Phase 1", "FE+BE", "POC dynamic; integrate with F8 engine"),
        (21, "A", "Collaboration & Approval", "Workflow lifecycle, per-field comments, @mentions, needs-you queue", "P1 — High", "Partial", 6, "F9, ClientSpace", "Phase 1", "FE+BE", "Workflow engine + notification system"),
        (22, "A", "Booklet Generation & Versioning", "Auto-generate booklet from F7 templates + CAP data, version, snapshot on sign-off", "P1 — High", "Partial", 5, "F7 template engine", "Phase 1", "BE", "PDF generation service"),
        (23, "A", "Client Sign-off (DocuSign)", "Route booklet via DocuSign, track status, advance lifecycle", "P1 — High", "POC Done", 3, "F2 DocuSign", "Phase 1", "BE", "POC state machine done; wire to real API"),
        (24, "A", "Ben Admin Handoff & Prism Write-back", "Structured payload, Ben Admin review, Prism API write-back, ClientSpace mirror", "P0 — Critical Path", "Partial", 8, "F2 Prism, CS", "Phase 1", "BE", "createBenefitPlanSetup integration"),

        # RENEWAL section
        ("RENEWAL — CAP Renewal Flow (R1-R9)", None),
        (25, "A", "Renewal Radar / Book of Business", "Queue of upcoming renewals from synced client records, urgency, owner, risk", "P0 — Critical Path", "POC Done", 2, "F2 Stream 2", "Phase 1", "FE", "POC has 8-client queue; swap to real data"),
        (26, "A", "Start Renewal & Pre-fill", "Per-client refresh -> pre-fill from prior-year CAP + Prism + plan-code crosswalk", "P0 — Critical Path", "POC Done", 4, "F2 Stream 2, F3", "Phase 1", "FE+BE", "Add real Prism refresh call"),
        (27, "A", "Document Updates as Deltas", "Upload -> extracted values as proposed deltas (accept/reject vs baseline)", "P1 — High", "POC Done", 4, "NB2 extraction", "Phase 1", "FE+BE", "Reuses NB2 extraction engine"),
        (28, "A", "Data-Currency & Version Drift", "Version comparison table, 'Bring Up to Date' with reviewable diff", "P0 — Critical Path", "POC Done", 3, "F2 drift detection", "Phase 1", "FE+BE", "POC has version table + rebase"),
        (29, "A", "Renewal Diff & Review", "Unified diff: client vs master changes, accept/flag, bulk approve", "P0 — Critical Path", "POC Done", 2, "R2, R4", "Phase 1", "FE", "POC fully interactive"),
        (30, "A", "Readiness & Approval", "Dynamic readiness gates, Fix navigation, role-gated approve", "P0 — Critical Path", "POC Done", 2, "F8, F9", "Phase 1", "FE+BE", "POC gates are dynamic"),
        (31, "A", "Booklet & Annual Contract Mgmt", "Booklet preview + generate, YoY sign-off history, route for signature", "P1 — High", "POC Done", 3, "F7, NB7", "Phase 1", "FE+BE", "Shares booklet engine with NB7"),
        (32, "A", "Client Sign-off", "DocuSign routing, shared infrastructure with NB8", "P1 — High", "POC Done", 1, "NB8", "Phase 1", "BE", "Reuses NB8 flow"),
        (33, "A", "Prism Write-back (Update)", "Structured update payload + downstream checklist + Prism API update", "P0 — Critical Path", "POC Done", 5, "NB9", "Phase 1", "BE", "Update operation needs plan-code migration"),

        # COPILOT section
        ("COPILOT — AI Assistant (CP1-CP7)", None),
        (34, "A", "Sidebar & Interaction Model", "Persistent sidebar, system/user/agent messages, actionable cards", "P0 — Critical Path", "POC Done", 1, "None", "Phase 1", "FE", "POC complete"),
        (35, "A", "Context Awareness & Grounding", "Context assembly per turn: CAP ID, section, role, pinned versions", "P1 — High", "Partial", 5, "LLM orchestration", "Phase 1", "BE+AI", "Context window management"),
        (36, "A", "Retrieve & Surface (Read Tools)", "MCP-based read tools across CAP, Foundation, Prism, CS, DocuSign", "P1 — High", "Partial", 5, "MCP servers", "Phase 1", "BE+AI", "Wire to MCP tool-calling"),
        (37, "A", "Act & Update (Write Tools)", "Before->after preview, Apply/Discard, field edits, modeling", "P1 — High", "POC Done", 4, "CP3, MCP", "Phase 1", "FE+BE+AI", "POC has proposed changes; wire to real writes"),
        (38, "A", "Explain & Reason", "Rate math walkthrough, blocker explanation, YoY diff analysis", "P1 — High", "Partial", 3, "CP2, CP3", "Phase 1", "AI", "LLM prompt engineering"),
        (39, "A", "Suggested Actions / Proactive", "Context-aware chips, proactive nudges", "P2 — Medium", "POC Done", 2, "CP2", "Phase 1", "FE+AI", "Enhance with state-aware nudges"),
        (40, "A", "Guardrails, Permissions & Audit", "RBAC on copilot, PII minimization, synced-field refusal, audit", "P0 — Critical Path", "POC Done", 3, "F9, MCP", "Phase 1", "BE", "POC has guardrails; harden for prod"),

        # INFRASTRUCTURE section
        ("INFRASTRUCTURE & CROSS-CUTTING (Blueprint Tab 3, 7, 8)", None),
        (41, "A", "Database & Data Layer", "PostgreSQL schema: Client, CAP, Product, Plan, PlanTier, UW, AdminConfig, TaxPlan, Document, AuditLog + Foundation entities", "P0 — Critical Path", "Not Started", 6, "None", "Phase 1", "BE", "Schema from ERD — straightforward migration from JSON"),
        (42, "A", "Authentication & SSO", "SAML SSO with G&A IdP, JWT sessions, role resolution", "P0 — Critical Path", "Not Started", 3, "G&A IdP access", "Phase 1", "BE", "Standard SAML integration"),
        (43, "A", "API Layer (Backend)", "Next.js API routes or FastAPI: CAP CRUD, Foundation CRUD, validation, rate calc, doc gen", "P0 — Critical Path", "Not Started", 10, "Database", "Phase 1", "BE", "Most logic exists in POC frontend; lift to API"),
        (44, "A", "MCP Server — cap-data-mcp", "8 tools + 6 resources wrapping platform data (Blueprint Phase 1)", "P1 — High", "Design Only", 4, "API layer", "Phase 1", "BE", "Architecture defined; standard MCP SDK"),
        (45, "A", "MCP Server — prismhr-mcp", "12 tools + 4 resources wrapping PrismHR APIs (Phase 2)", "P1 — High", "Design Only", 6, "Prism API access", "Phase 1", "BE", "Critical for Copilot + write-back"),
        (46, "A", "MCP Server — CS/DocuSign/WS", "Additional MCP servers for workflow, e-sign, enrollment (Phase 3-4)", "P2 — Medium", "Design Only", 5, "Respective APIs", "Phase 1", "BE", "Enables cross-system Copilot actions"),
        (47, "A", "Audit Trail System", "Field-change logging: who, when, what, old->new. SOX controls.", "P0 — Critical Path", "POC Done", 3, "Database", "Phase 1", "BE", "POC has real-time store; add DB persistence"),
        (48, "A", "Versioning & Snapshot Engine", "Effective-dating for G&A config, snapshot-on-finalize for signed CAPs", "P0 — Critical Path", "Not Started", 5, "Database", "Phase 1", "BE", "Version table pattern"),
        (49, "A", "Notification & Alerting", "In-app + email: renewal reminders, approval requests, SLA escalation", "P2 — Medium", "Not Started", 4, "Email service", "Phase 1", "BE", "Standard notification service"),
        (50, "A", "Data Governance & PII", "Field-level classification, encryption at rest + transit, HIPAA BAA", "P1 — High", "Design Only", 4, "Cloud provider", "Phase 1", "DevOps", "Blueprint Tab 7 classification scheme"),
        (51, "A", "CI/CD & Cloud Deployment", "Docker, staging + prod, automated testing, blue-green deploy", "P0 — Critical Path", "Not Started", 4, "Cloud account", "Phase 1", "DevOps", "Standard pipeline"),
        (52, "A", "Reporting & Analytics", "CAP metrics, renewal pipeline, throughput, contribution trends", "P3 — Nice-to-have", "Not Started", 5, "Database, API", "Phase 1", "FE+BE", "Post-launch enhancement"),

        # ─── Initiative B ───
        ("INITIATIVE B — AI Underwriting Assistant", INIT_B_BG),
        (53, "B", "AI Document Extraction Engine", "Core extraction pipeline: ingest PDFs, SBCs, carrier invoices; OCR + layout analysis; output structured plan data with field-level confidence scores", "P0 — Critical Path", "Not Started", 10, "Document storage, OCR service", "Phase 2", "AI+BE", "Handles 10,426 annual quotes; target 20-30 min savings per case"),
        (54, "B", "SBC & Invoice Parser", "Specialized parsers for Summary of Benefits & Coverage documents and carrier invoices; extract plan design fields, deductibles, copays, coinsurance, rate tiers", "P0 — Critical Path", "Not Started", 7, "B-53 Extraction Engine", "Phase 2", "AI", "SBC format is ACA-standardized; invoice formats vary by carrier"),
        (55, "B", "Comparison Builder", "Auto-populate underwriting comparison spreadsheet from extracted data; side-by-side carrier comparison with rate analysis and concession history", "P0 — Critical Path", "Not Started", 6, "B-54 SBC Parser", "Phase 2", "FE+BE", "Replaces manual 20-30 min scrubbing per quote"),
        (56, "B", "Pre-Submission Completeness Gate", "Validation engine that blocks incomplete submissions from entering the underwriting queue; actionable error messages listing missing fields", "P1 — High", "Not Started", 5, "B-53 Extraction Engine", "Phase 2", "BE", "80-90% of submissions currently arrive incomplete"),
        (57, "B", "Hero/Gradient Integration", "Push extracted census + plan data to Hero/Gradient risk scoring platforms; pull risk scores back into comparison workflow", "P1 — High", "Not Started", 5, "B-55 Comparison Builder, Hero/Gradient API", "Phase 2", "BE", "Existing external platforms; API integration required"),
        (58, "B", "Extraction Confidence Scoring & Review UI", "Dashboard showing extraction confidence per field; human review queue for low-confidence fields; bulk approve for high-confidence extractions", "P1 — High", "Not Started", 4, "B-53 Extraction Engine", "Phase 2", "FE+AI", "Specialist reviews and confirms; system never auto-commits"),
        (59, "B", "Age-Banded Rate Calculator", "Automated age-banded rate calculation for large group quotes; tier mapping and composite rate derivation from extracted census data", "P2 — Medium", "Not Started", 3, "B-54 SBC Parser, F4 Pricing Stack", "Phase 2", "BE", "Large group quotes currently extend scrubbing time significantly"),

        # ─── Initiative C ───
        ("INITIATIVE C — Carrier Portal Automation", INIT_C_BG),
        (60, "C", "Carrier Portal Inventory & Mapping", "Comprehensive inventory of all open-market carrier portals; field mapping from Prism enrollment data to each carrier's portal fields; API vs RPA classification per carrier", "P0 — Critical Path", "Not Started", 5, "Carrier portal access", "Phase 2", "BA+BE", "Classification determines API vs RPA approach per carrier"),
        (61, "C", "API Integration — Master Carriers", "Direct API integration for carriers with published APIs (e.g., Guardian); Prism-to-carrier enrollment data push with confirmation capture", "P0 — Critical Path", "Not Started", 8, "C-60 Portal Inventory, Carrier API access", "Phase 2", "BE", "Start with highest-volume carriers; API preferred over RPA"),
        (62, "C", "RPA Framework for Non-API Portals", "Browser automation framework for carrier portals without API; form-fill from Prism data, screenshot capture, error handling", "P1 — High", "Not Started", 7, "C-60 Portal Inventory", "Phase 2", "BE+DevOps", "Eliminates 13-17 min manual entry per enrollment"),
        (63, "C", "Enrollment Data Mapping from Prism", "Standardized data extraction from Prism enrollment records; transform and validate against carrier-specific field requirements before submission", "P0 — Critical Path", "Not Started", 5, "F2 PrismHR Connector", "Phase 2", "BE", "Prism read integration already built in Phase 1"),
        (64, "C", "Automated Enrollment Audit & Reconciliation", "Post-enrollment verification: compare submitted data against carrier confirmation; flag discrepancies; maintain compliance audit trail in ClientSpace", "P1 — High", "Not Started", 5, "C-61/C-62 Carrier Integration", "Phase 2", "BE", "SOX-adjacent — coverage gap risk requires audit step"),
        (65, "C", "Exception Routing & Specialist Review", "Route enrollment exceptions to assigned Benefits Specialist with field-level context; specialist retains audit review authority", "P1 — High", "Not Started", 5, "C-64 Audit, ClientSpace", "Phase 2", "FE+BE", "Specialist audit review step always retained per PRD"),

        # ─── Initiative D ───
        ("INITIATIVE D — Service Automation (Ciklum Expansion)", INIT_D_BG),
        (66, "D", "Automated Case Creation from Intake", "Integration with Ciklum to auto-create ClientSpace cases from intake channels (phone, email, portal); eliminate manual triage logging", "P1 — High", "Not Started", 7, "Ciklum platform, ClientSpace API", "Phase 3", "BE", "Benefits Specialists spend 35% of time on WSE/client issue resolution"),
        (67, "D", "Self-Service Routing for ID Cards & FAQs", "Ciklum self-service layer routes low-complexity inquiries (ID cards, coverage verification) to carrier apps with guided prompts; complex cases to Specialist", "P1 — High", "Not Started", 6, "Ciklum platform, Carrier apps", "Phase 3", "BE+FE", "2-6 ID card requests/day; 30-50 during peak OE"),
        (68, "D", "AI After-Call Summarization", "Auto-generate case notes from call transcripts; populate ClientSpace case records; flag follow-up actions", "P2 — Medium", "Not Started", 6, "Ciklum platform, ClientSpace API", "Phase 3", "AI+BE", "Eliminates manual post-call documentation"),
        (69, "D", "ClientSpace Integration Layer", "Bi-directional sync between Ciklum and ClientSpace for case status, assignment, and escalation; unified view across both platforms", "P1 — High", "Not Started", 6, "ClientSpace API, Ciklum API", "Phase 3", "BE", "Critical for maintaining single source of truth for case management"),

        # ─── Initiative E ───
        ("INITIATIVE E — Payroll Deduction Audit (Sidekick Expansion)", INIT_E_BG),
        (70, "E", "Consolidated Prism Report Generation", "Sidekick integration to auto-pull and merge 4+ Prism reports (deduction register, plan rates, prior period data) into single consolidated view", "P0 — Critical Path", "Not Started", 6, "Sidekick platform, Prism API", "Phase 3", "BE", "Currently 20-60 min/client pulling and combining reports manually"),
        (71, "E", "CAP-to-Deduction Comparison Engine", "Automated tier-by-tier comparison of Prism deductions against CAP rates; highlight mismatches with drill-down to employee level", "P0 — Critical Path", "Not Started", 5, "E-70 Consolidated Report, CAP Platform (Init A)", "Phase 3", "BE", "CAP Cloud Tool provides structured rate data as comparison baseline"),
        (72, "E", "HSA Employer Match Auto-Calculation", "Automated calculation of HSA employer match amounts from plan enrollment data and CAP specs; validate against Prism entries per employee", "P1 — High", "Not Started", 4, "E-71 Comparison Engine, CAP Platform", "Phase 3", "BE", "Currently entered manually per employee with no automated validation"),
        (73, "E", "Exception-Based Variance Flagging", "Surface only mismatched records for Specialist review; routine matches confirmed in bulk; exception routing to ClientSpace with variance detail", "P1 — High", "Not Started", 5, "E-71 Comparison Engine, ClientSpace", "Phase 3", "FE+BE", "Specialists review exceptions only; 23% of time currently on payroll reviews"),

        # ─── Initiative F ───
        ("INITIATIVE F — Benefits Reconciliation Tool Rollout", INIT_F_BG),
        (74, "F", "Beacon UAT Completion", "Complete user acceptance testing of Beacon reconciliation platform; document and resolve all blocking defects; validate employee-level invoice comparison accuracy", "P0 — Critical Path", "Not Started", 6, "Beacon platform access", "Phase 3", "QA+BE", "Beacon partially deployed but unreliable; Excel fallback currently required"),
        (75, "F", "Production Stabilization & Monitoring", "Production hardening: performance tuning, error handling, alerting dashboards; eliminate monthly Excel fallback dependency", "P0 — Critical Path", "Not Started", 5, "F-74 UAT Complete", "Phase 3", "DevOps+BE", "Recon Specialists spend 97% of time on reconciliation"),
        (76, "F", "Payment Routing Automation", "Automate carrier payment routing logic across 8 different payment methods; integrate with reconciliation output for exception-based payment approval", "P1 — High", "Not Started", 4, "F-75 Production Stable", "Phase 3", "BE", "8 different carrier payment methods currently require manual routing"),
    ]

    r = 5
    section_colors = {
        "FOUNDATION": SECTION_BG,
        "NEW BUSINESS": SECTION_BG,
        "RENEWAL": SECTION_BG,
        "COPILOT": SECTION_BG,
        "INFRASTRUCTURE": SECTION_BG,
    }

    for item in init_a_data:
        if isinstance(item[1], str) and item[1] is None or (len(item) == 2):
            # Section header
            label = item[0]
            bg_color = item[1] if item[1] else SECTION_BG
            fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            cell = ws.cell(row=r, column=1, value=label)
            cell.font = section_font
            cell.fill = fill
            cell.border = thin_border
            for c in range(2, NUM_COLS + 1):
                cl = ws.cell(row=r, column=c)
                cl.fill = fill
                cl.border = thin_border
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NUM_COLS)
            r += 1
        else:
            # Data row
            apply_data_row(ws, r, list(item))
            r += 1

    # Freeze panes at header
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:K{r-1}"

    return ws


# =============================================================================
# SHEET 2: Sprint Plan
# =============================================================================
def build_sprint_plan(wb):
    ws = wb.create_sheet("Sprint Plan")
    NUM_COLS = 7

    set_col_widths(ws, [8, 12, 28, 55, 16, 40, 35])

    # Title
    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NUM_COLS)
    ws.cell(r, 1, "Sprint Roadmap — 11 Sprints (22 Weeks) · 3 Phases · 6 Initiatives").font = title_font

    r = 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NUM_COLS)
    ws.cell(r, 1, "Accelerated timeline leveraging comprehensive POC codebase as production scaffold and proven business logic from POC validation").font = subtitle_font

    r = 3  # blank

    # Header
    r = 4
    headers = ["Sprint", "Duration", "Theme", "Work Items (Ref #)", "Team Focus", "Key Deliverable", "Dependencies / Risks"]
    apply_header_row(ws, r, headers)

    # Phase 1 section header
    r = 5
    apply_section_row(ws, r, "PHASE 1 — Initiative A: CAP Platform (Sprints 1-6, 12 Weeks)", NUM_COLS)

    sprints = [
        ("S1", "2 weeks", "Foundation + DB + Auth",
         "#41 DB schema, #42 SSO, #43 API scaffold, #51 CI/CD, #1 Admin shell, #2 PrismHR connector, #6 Sync Stream 1",
         "BE + DevOps",
         "DB live, SAML login working, first Prism sync running, API skeleton deployed",
         "Prism API service account by day 1"),
        ("S2", "2 weeks", "Full Foundation + NB/Renewal Core",
         "#7-11 Streams+Plans+Pricing+Params+Vocab, #13-14 Validation+Roles, #47-48 Audit+Versioning, #16-20 NB1-NB5, #25-30 R1-R6",
         "FE + BE",
         "All F1-F10 live with real data. NB wizard + Renewal flow saving to DB. Dynamic validation working.",
         "Heavy sprint — POC code lifts directly to prod"),
        ("S3", "2 weeks", "Documents + E-Sign + Copilot Start",
         "#12 Templates, #15 UW Intake, #17 AI Extraction, #21-23 NB6-NB8, #27 R3 Doc Deltas, #31-32 R7-R8, #3-4 CS+DocuSign, #34+40 CP1+CP7, #44 cap-data-mcp, #49 Notifications",
         "Full team",
         "Booklet generation, DocuSign integration, Copilot sidebar with MCP tools, notification system",
         "Template engine + DocuSign API key needed"),
        ("S4", "2 weeks", "Write-back + Copilot Full + MCP",
         "#24 NB9 Prism write-back, #33 R9 update write-back, #35-39 CP2-CP6, #45-46 MCP servers, #5 WorkSight, #50 Data governance, #52 Reporting",
         "BE + AI",
         "End-to-end write-back to Prism. Full Copilot with LLM + MCP tool-calling. All integrations live.",
         "LLM provider access + Prism write perms"),
        ("S5", "2 weeks", "Hardening + Security + Performance",
         "E2E testing, rate calculation parallel-run vs Excel, HIPAA compliance audit, performance tuning, edge cases (multi-entity, open-market, age-banded rates)",
         "QA + DevOps",
         "Rate parity verified. Security review passed. Load-tested for 50 concurrent users.",
         "Need 10 real client CAPs for parallel-run"),
        ("S6", "2 weeks", "UAT + Pilot + Launch",
         "User acceptance testing with 5 pilot AMs, data migration from 20 Excel CAPs, training sessions, go-live checklist, monitoring dashboards",
         "Full team + G&A",
         "Production launch with pilot group. Excel-to-platform migration proven.",
         "AM availability for UAT. Change management plan."),
    ]

    r = 6
    for sprint in sprints:
        apply_data_row(ws, r, list(sprint))
        r += 1

    # Phase 2 section header
    apply_section_row(ws, r, "PHASE 2 — Initiatives B + C: AI Underwriting + Carrier Portal (Sprints 7-9, 6 Weeks)", NUM_COLS)
    r += 1

    phase2_sprints = [
        ("S7", "2 weeks", "AI Extraction Pipeline + Portal Inventory",
         "#53 AI Extraction Engine, #54 SBC/Invoice Parser, #56 Completeness Gate, #60 Carrier Portal Inventory, #63 Prism Enrollment Data Mapping",
         "AI + BE",
         "Core extraction pipeline processing real documents. Carrier portal inventory complete with API/RPA classification.",
         "Sample SBCs and invoices for training. Carrier portal access credentials."),
        ("S8", "2 weeks", "Comparison Builder + Carrier API Integration",
         "#55 Comparison Builder, #57 Hero/Gradient Integration, #58 Confidence Scoring UI, #61 API Integration (Master Carriers), #62 RPA Framework",
         "Full team",
         "End-to-end underwriting comparison auto-populated from documents. First carrier API enrollments automated.",
         "Hero/Gradient API access. Guardian API credentials."),
        ("S9", "2 weeks", "Polish + Remaining Carriers + Validation",
         "#59 Age-Banded Calculator, #64 Enrollment Audit/Recon, #65 Exception Routing, Phase 2 E2E testing and hardening",
         "BE + QA",
         "All Init B+C features in production. Enrollment audit trail complete. Exception routing to Specialists.",
         "Carrier confirmation of API/RPA compatibility. Specialist UAT for exception workflow."),
    ]
    for sprint in phase2_sprints:
        apply_data_row(ws, r, list(sprint))
        r += 1

    # Phase 3 section header
    apply_section_row(ws, r, "PHASE 3 — Initiatives D + E + F: Platform Expansions (Sprints 10-11, 4 Weeks)", NUM_COLS)
    r += 1

    phase3_sprints = [
        ("S10", "2 weeks", "Ciklum + Sidekick Integration + Beacon UAT",
         "#66 Auto Case Creation, #67 Self-Service Routing, #70 Consolidated Prism Reports, #71 CAP-to-Deduction Comparison, #74 Beacon UAT Completion",
         "BE + QA",
         "Ciklum case creation live. Sidekick deduction audit consolidated. Beacon UAT defects resolved.",
         "Ciklum and Sidekick vendor coordination. Beacon platform access."),
        ("S11", "2 weeks", "Remaining Integrations + Stabilization + Launch",
         "#68 AI Summarization, #69 ClientSpace Integration, #72 HSA Auto-Calc, #73 Variance Flagging, #75 Beacon Production Stabilization, #76 Payment Routing",
         "Full team",
         "All 6 initiatives in production. Full platform operational. Monitoring dashboards for all modules.",
         "Vendor SLAs for Ciklum/Sidekick. Beacon production environment access."),
    ]
    for sprint in phase3_sprints:
        apply_data_row(ws, r, list(sprint))
        r += 1

    # Summary row
    r += 1
    summary_data = ["TOTAL", "22 weeks", "Full G&A Benefits Platform",
                    "76 work items across 6 initiatives",
                    "~9 peak FTEs",
                    "Cloud CAP + AI UW + Carrier Automation + Service/Deduction/Recon platform expansions",
                    "Phased delivery reduces risk; each phase delivers standalone value"]
    apply_data_row(ws, r, summary_data, font=bold_font, fill=total_fill)

    ws.freeze_panes = "A5"
    return ws


# =============================================================================
# SHEET 3: Effort Summary
# =============================================================================
def build_effort_summary(wb):
    ws = wb.create_sheet("Effort Summary")
    NUM_COLS = 10

    set_col_widths(ws, [32, 14, 10, 10, 12, 12, 10, 12, 16, 16])

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NUM_COLS)
    ws.cell(r, 1, "Full Platform Effort Summary — 6 Initiatives · 22-Week Phased Delivery").font = title_font

    r = 2  # blank

    r = 3
    headers = ["Category", "Module Count", "P0 Items", "P1 Items", "P2/P3 Items",
               "POC Done", "Partial", "Not Started", "Est. Prod Days", "Est. FTEs (22 wks)"]
    apply_header_row(ws, r, headers)

    # Phase 1 section
    r = 4
    apply_section_row(ws, r, "PHASE 1 — Initiative A: CAP Platform (12 Weeks)", NUM_COLS)

    rows = [
        ("Init A: Foundation (F1-F10)",  15, 6, 7,  2,  1, 12, 2,  65, "1.8 FTE"),
        ("Init A: New Business (NB1-NB9)", 9, 5, 4,  0,  5,  3, 1,  38, "1.1 FTE"),
        ("Init A: Renewal (R1-R9)",        9, 6, 3,  0,  8,  1, 0,  26, "0.7 FTE"),
        ("Init A: Copilot (CP1-CP7)",      7, 2, 4,  1,  3,  3, 1,  23, "0.6 FTE"),
        ("Init A: Infrastructure",        12, 5, 3,  4,  1,  0, 8,  59, "1.6 FTE"),
    ]

    r = 5
    for row in rows:
        apply_data_row(ws, r, list(row))
        r += 1

    # Phase 1 subtotal
    apply_data_row(ws, r, ["Phase 1 Subtotal (Init A)", 52, 24, 21, 7, 18, 19, 12, 211, "~6 FTEs x 12 wks"],
                   font=bold_font, fill=PatternFill(start_color=HIGHLIGHT_BLUE, end_color=HIGHLIGHT_BLUE, fill_type="solid"))
    r += 1

    # Phase 2 section
    r += 1
    apply_section_row(ws, r, "PHASE 2 — Initiatives B + C: AI Underwriting + Carrier Portal (6 Weeks)", NUM_COLS)
    r += 1

    phase2_rows = [
        ("Init B: AI Underwriting Assistant", 7, 3, 3, 1, 0, 0, 7, 40, "1.5 FTE"),
        ("Init C: Carrier Portal Automation", 6, 3, 3, 0, 0, 0, 6, 35, "1.3 FTE"),
    ]
    for row in phase2_rows:
        apply_data_row(ws, r, list(row))
        r += 1

    apply_data_row(ws, r, ["Phase 2 Subtotal (Init B+C)", 13, 6, 6, 1, 0, 0, 13, 75, "~8 FTEs x 6 wks"],
                   font=bold_font, fill=PatternFill(start_color=HIGHLIGHT_BLUE, end_color=HIGHLIGHT_BLUE, fill_type="solid"))
    r += 1

    # Phase 3 section
    r += 1
    apply_section_row(ws, r, "PHASE 3 — Initiatives D + E + F: Platform Expansions (4 Weeks)", NUM_COLS)
    r += 1

    phase3_rows = [
        ("Init D: Service Automation (Ciklum)", 4, 0, 3, 1, 0, 0, 4, 25, "1.4 FTE"),
        ("Init E: Payroll Deduction Audit (Sidekick)", 4, 2, 2, 0, 0, 0, 4, 20, "1.1 FTE"),
        ("Init F: Reconciliation Tool (Beacon)", 3, 2, 1, 0, 0, 0, 3, 15, "0.8 FTE"),
    ]
    for row in phase3_rows:
        apply_data_row(ws, r, list(row))
        r += 1

    apply_data_row(ws, r, ["Phase 3 Subtotal (Init D+E+F)", 11, 4, 6, 1, 0, 0, 11, 60, "~9 FTEs x 4 wks"],
                   font=bold_font, fill=PatternFill(start_color=HIGHLIGHT_BLUE, end_color=HIGHLIGHT_BLUE, fill_type="solid"))
    r += 1

    # Grand Total
    r += 1
    apply_data_row(ws, r, ["GRAND TOTAL — ALL INITIATIVES", 76, 34, 33, 9, 18, 19, 36, 346, "22 weeks total"],
                   font=Font(name="Calibri", size=11, bold=True, color=WHITE),
                   fill=PatternFill(start_color=DARK_HEADER_BG, end_color=DARK_HEADER_BG, fill_type="solid"),
                   align=Alignment(horizontal="center", vertical="center", wrap_text=True))
    r += 1

    # Accelerator notes
    r += 1
    notes = [
        "Key accelerators (Phase 1 — Initiative A):",
        "1. Comprehensive POC serves as production scaffold — 80% of UI components lift directly with backend wiring",
        "2. POC validated all business logic (rate formulas, validation rules, contribution strategies) — no re-discovery needed",
        "3. MCP architecture standardizes all integrations — no custom wiring per feature",
        "4. Dynamic validation, audit trail, and role system already proven in POC — prod adds persistence layer only",
        "5. Renewal flow (R1-R9) has 8/9 modules POC Done — fastest path to value",
        "",
        "Key accelerators (Phase 2 — Initiatives B+C):",
        "1. Initiative A infrastructure (Prism connector, data layer, auth) reused directly — no duplicate integration work",
        "2. Extraction pipeline architecture proven during Init A (NB2/F10) — scales to full underwriting scope",
        "3. Carrier portal field mapping leverages Prism data layer built in Phase 1",
        "",
        "Key accelerators (Phase 3 — Initiatives D+E+F):",
        "1. Ciklum and Sidekick are vendor platform expansions — our scope is integration, not core platform build",
        "2. Beacon UAT is testing and stabilization, not new development",
        "3. ClientSpace integration pattern established in Phase 1 — reused for D, E, and F",
    ]
    for note in notes:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NUM_COLS)
        cell = ws.cell(r, 1, note)
        if note.startswith("Key accelerators"):
            cell.font = bold_font
        else:
            cell.font = note_font
        r += 1

    ws.freeze_panes = "A4"
    return ws


# =============================================================================
# SHEET 4: Risk Register
# =============================================================================
def build_risk_register(wb):
    ws = wb.create_sheet("Risk Register")
    NUM_COLS = 8

    set_col_widths(ws, [5, 12, 30, 10, 12, 10, 45, 10])

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NUM_COLS)
    ws.cell(r, 1, "Full Platform Migration — Risk Register").font = title_font

    r = 2  # blank

    r = 3
    headers = ["#", "Initiative", "Risk", "Impact", "Likelihood", "Severity", "Mitigation", "Owner"]
    apply_header_row(ws, r, headers)

    risks = [
        # Phase 1 risks (Init A)
        ("PHASE 1 RISKS — Initiative A: CAP Platform", None),
        (1,  "A", "PrismHR API access not provisioned in time", "High", "Medium", "High",
         "Request in week 0. POC abstraction layer allows simulated data until live access arrives — zero blocking.", "PM"),
        (2,  "A", "AI extraction accuracy below 90% confidence", "Medium", "Medium", "Medium",
         "Human-confirmation workflow (NB2/NB3) is built and proven. System augments, never auto-commits.", "AI Lead"),
        (3,  "A", "HIPAA compliance gaps", "High", "Low", "Medium",
         "SOC 2 + BAA cloud. Field-level PII classification from Tab 7 already designed. Security review in S5.", "DevOps"),
        (4,  "A", "AM adoption resistance (27 AMs)", "High", "High", "Critical",
         "Pilot with 5 AMs in S6. Copilot reduces learning curve dramatically. Side-by-side Excel comparison.", "PM"),
        (5,  "A", "Rate calc discrepancy vs Excel", "Critical", "Medium", "Critical",
         "Parallel-run in S5: both systems compute rates for 10 clients, compare to $0.01. Formula proven in POC.", "BE Lead"),
        (6,  "A", "Compressed timeline — quality risk", "Medium", "Medium", "Medium",
         "POC de-risks significantly: all UX validated, all business logic verified, all flows tested. Prod is wiring, not invention.", "Tech Lead"),
        (7,  "A", "Open-market plans (no API/EDI)", "Medium", "High", "High",
         "Manual rate entry as primary path (matches current process). AI extraction as accelerator.", "FE Lead"),
        (8,  "A", "Data migration from 100+ Excel CAPs", "High", "High", "High",
         "Build parser targeting Tab 1 structure. Batch import with validation. Keep Excel as read-only archive.", "BE + Data"),

        # Phase 2 risks (Init B + C)
        ("PHASE 2 RISKS — Initiatives B + C: AI Underwriting + Carrier Portal", None),
        (9,  "B", "Document format variability across carriers", "High", "High", "High",
         "Start with ACA-standardized SBC format (consistent structure). Invoice parsers prioritized by carrier volume. Confidence scoring ensures low-confidence extractions route to human review.", "AI Lead"),
        (10, "B", "Extraction accuracy insufficient for production use", "High", "Medium", "High",
         "Human-in-the-loop design from day 1 — extraction is always review-and-confirm, never auto-commit. Confidence threshold tuned per field type.", "AI Lead"),
        (11, "B", "Hero/Gradient API access or compatibility issues", "Medium", "Medium", "Medium",
         "Hero/Gradient are existing platforms already in use. API documentation review in Sprint 7; fallback is manual entry into existing platforms.", "BE Lead"),
        (12, "C", "Carrier portal changes break RPA scripts", "High", "High", "Critical",
         "API integration preferred over RPA where possible. RPA scripts include visual validation and screenshot capture. Monitoring alerts on failure. Specialist manual fallback always available.", "DevOps"),
        (13, "C", "Carrier reluctance to provide API access", "Medium", "High", "High",
         "Start with carriers who have published APIs (e.g., Guardian). RPA as bridge for non-API carriers. Carrier relationship managed by G&A Benefits team.", "PM"),
        (14, "C", "Enrollment data mapping errors create coverage gaps", "Critical", "Medium", "Critical",
         "Dual-confirmation: automated enrollment verified against carrier confirmation response. Specialist audit review retained on every enrollment. SOX-compliant audit trail.", "BE Lead"),

        # Phase 3 risks (Init D + E + F)
        ("PHASE 3 RISKS — Initiatives D + E + F: Platform Expansions", None),
        (15, "D", "Ciklum vendor delivery timeline misalignment", "Medium", "Medium", "Medium",
         "Our scope is integration layer only — Ciklum owns platform expansion. Parallel workstreams with clear interface contracts. Fallback: current manual process continues.", "PM"),
        (16, "D", "Self-service deflection rate lower than expected", "Medium", "High", "Medium",
         "Phased rollout: start with ID card requests (highest volume, simplest routing). Measure deflection rate before expanding to other inquiry types.", "BA"),
        (17, "E", "Sidekick platform limitations for Prism integration", "Medium", "Medium", "Medium",
         "Sidekick is existing vendor with established Prism connectivity. Integration scope limited to consolidated report push and variance flagging. Fallback: current Prism report process.", "BE Lead"),
        (18, "E", "CAP rate data not structured for automated comparison", "Low", "Low", "Low",
         "Init A (Phase 1) delivers structured CAP data in the cloud platform — this is the primary enabler. Comparison engine reads directly from Init A data layer.", "BE Lead"),
        (19, "F", "Beacon UAT reveals fundamental platform defects", "High", "Medium", "High",
         "Scope is UAT completion and stabilization, not new development. If defects are severe, escalate to Beacon vendor. Excel fallback remains operational during stabilization.", "QA Lead"),
        (20, "F", "Recon Specialist resistance to Beacon adoption", "Medium", "Medium", "Medium",
         "Beacon is already partially deployed — team is familiar with the tool. UAT involves Recon Specialists directly. Excel fallback eliminated only after production stability confirmed.", "PM"),
    ]

    r = 4
    for item in risks:
        if isinstance(item, tuple) and len(item) == 2:
            # Section header
            apply_section_row(ws, r, item[0], NUM_COLS)
            r += 1
        else:
            apply_data_row(ws, r, list(item))
            r += 1

    ws.freeze_panes = "A4"
    return ws


# =============================================================================
# SHEET 5: Team & Resources
# =============================================================================
def build_team_resources(wb):
    ws = wb.create_sheet("Team & Resources")
    NUM_COLS = 5

    set_col_widths(ws, [28, 10, 22, 50, 40])

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NUM_COLS)
    ws.cell(r, 1, "Team Composition — Full Platform, Phased Delivery (22 Weeks)").font = title_font

    r = 2  # blank

    # Phase 1 Team
    r = 3
    apply_section_row(ws, r, "PHASE 1 — Initiative A: CAP Platform (Sprints 1-6, 12 Weeks) — 6 FTEs", NUM_COLS)

    r = 4
    headers = ["Role", "Count", "Sprint Allocation", "Key Responsibilities", "Skills Required"]
    apply_header_row(ws, r, headers)

    phase1_team = [
        ("Senior Frontend Engineer", 1, "S1-S6 (full)",
         "React/Next.js components (lift from POC), Zustand-to-backend migration, Copilot UI, accessibility",
         "React, TypeScript, Next.js App Router, inline styles, Zustand"),
        ("Senior Backend Engineer", 2, "S1-S6 (full)",
         "API design, PostgreSQL, PrismHR/CS/DocuSign integration, MCP servers, rate engine, validation engine",
         "Python/Node.js, PostgreSQL, REST APIs, MCP SDK, PrismHR API"),
        ("AI/ML Engineer", 1, "S2-S6",
         "Document extraction pipeline, LLM orchestration for Copilot, MCP tool-calling, prompt engineering",
         "LLM APIs, document extraction, MCP SDK, prompt engineering"),
        ("DevOps Engineer", 1, "S1, S4-S6",
         "CI/CD, Docker, cloud infra, monitoring, HIPAA compliance, staging/prod",
         "AWS/Azure, Docker, Terraform, SOC 2 compliance"),
        ("Product/Project Manager", 1, "S1-S6 (full)",
         "Sprint planning, G&A stakeholder mgmt, SME coordination, UAT, change management",
         "PEO/benefits domain, Agile delivery"),
    ]

    r = 5
    for row in phase1_team:
        apply_data_row(ws, r, list(row))
        r += 1

    apply_data_row(ws, r, ["Phase 1 Team Total", "~6 FTEs", "12 weeks", "~211 person-days of effort (Init A)", ""],
                   font=bold_font, fill=total_fill)
    r += 1

    # Phase 2 Team (add 2)
    r += 1
    apply_section_row(ws, r, "PHASE 2 — Initiatives B + C (Sprints 7-9, 6 Weeks) — Add 2 FTEs = 8 Total", NUM_COLS)
    r += 1
    apply_header_row(ws, r, headers)
    r += 1

    phase2_adds = [
        ("Phase 1 team continues", "6", "S7-S9 (all continuing)",
         "Core team from Phase 1 continues; BE engineers shift to carrier API integration; AI engineer shifts to extraction pipeline",
         "Same as Phase 1"),
        ("Senior AI Engineer (NEW)", 1, "S7-S9",
         "Production extraction pipeline: OCR, SBC parsing, invoice parsing, confidence scoring, age-banded rate handling",
         "Document AI, OCR pipelines, ML model deployment, PDF processing"),
        ("Integration/RPA Engineer (NEW)", 1, "S7-S9",
         "Carrier portal automation: API integration for supported carriers, RPA framework for non-API portals, enrollment data mapping",
         "API integration, browser automation (Selenium/Playwright), Prism data model, carrier portal experience"),
    ]
    for row in phase2_adds:
        apply_data_row(ws, r, list(row))
        r += 1

    apply_data_row(ws, r, ["Phase 2 Team Total", "~8 FTEs", "6 weeks", "~75 person-days of effort (Init B+C)", ""],
                   font=bold_font, fill=total_fill)
    r += 1

    # Phase 3 Team (add 1)
    r += 1
    apply_section_row(ws, r, "PHASE 3 — Initiatives D + E + F (Sprints 10-11, 4 Weeks) — Add 1 FTE = 9 Total", NUM_COLS)
    r += 1
    apply_header_row(ws, r, headers)
    r += 1

    phase3_adds = [
        ("Phase 2 team continues", "8", "S10-S11 (all continuing)",
         "Core team continues; AI/RPA engineers shift to Ciklum/Sidekick/Beacon integration and stabilization",
         "Same as Phase 2"),
        ("QA/Stabilization Engineer (NEW)", 1, "S10-S11",
         "Beacon UAT completion, production stabilization testing, cross-initiative regression testing, payment routing validation",
         "QA automation, reconciliation domain, UAT methodology, production monitoring"),
    ]
    for row in phase3_adds:
        apply_data_row(ws, r, list(row))
        r += 1

    apply_data_row(ws, r, ["Phase 3 Team Total", "~9 FTEs", "4 weeks", "~60 person-days of effort (Init D+E+F)", ""],
                   font=bold_font, fill=total_fill)
    r += 1

    # G&A-side resources
    r += 1
    apply_section_row(ws, r, "G&A-SIDE RESOURCES (Part-Time, All Phases)", NUM_COLS)
    r += 1
    apply_header_row(ws, r, headers)
    r += 1

    ga_resources = [
        ("Benefits SME (G&A)", 1, "S1-S6 (advisory, 50%)",
         "Domain validation, rate formula verification, workflow confirmation, UAT",
         "G&A CAP process, Prism knowledge"),
        ("UW SME (G&A)", 1, "S7-S9 (advisory, 25%)",
         "Underwriting workflow validation, extraction accuracy review, comparison template confirmation",
         "G&A underwriting process, Hero/Gradient"),
        ("Ben Admin SME (G&A)", 1, "S7-S11 (advisory, 25%)",
         "Carrier enrollment validation, payroll deduction audit confirmation, Beacon UAT participation",
         "Ben Admin processes, carrier portal experience"),
        ("IT/Systems Contact (G&A)", 1, "S1-S11 (as needed)",
         "Prism API provisioning, SSO configuration, carrier portal credentials, environment access",
         "G&A IT systems, Prism admin"),
    ]
    for row in ga_resources:
        apply_data_row(ws, r, list(row))
        r += 1

    # Why this works
    r += 2
    notes = [
        "Why this lean team works across 22 weeks:",
        "- POC eliminates discovery/design phase for Init A — team starts from validated, working code on day 1",
        "- 18 of 52 Init A work items are already POC Done — these need backend wiring only, not UI development",
        "- Phased delivery means team ramps gradually: 6 FTEs (Phase 1) -> 8 FTEs (Phase 2) -> 9 FTEs (Phase 3)",
        "- Phase 1 infrastructure (Prism connector, data layer, auth, MCP) is directly reused by Phases 2 and 3",
        "- Initiatives D, E, F are vendor platform expansions — our scope is integration, not core platform build",
        "- Compressed 22-week timeline vs typical 40+ weeks for greenfield — POC de-risked all business logic for Init A",
    ]
    for note in notes:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NUM_COLS)
        cell = ws.cell(r, 1, note)
        if note.startswith("Why"):
            cell.font = bold_font
        else:
            cell.font = note_font
        r += 1

    ws.freeze_panes = "A4"
    return ws


# =============================================================================
# SHEET 6: Business Impact
# =============================================================================
def build_business_impact(wb):
    ws = wb.create_sheet("Business Impact")

    # ── Section 1: Activity Baseline Table ──
    NUM_COLS = 9

    set_col_widths(ws, [32, 14, 14, 14, 14, 18, 14, 14, 14])

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NUM_COLS)
    ws.cell(r, 1, "Business Impact — G&A Beneficial Activity Baseline & Automation Yield").font = title_font

    r = 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NUM_COLS)
    ws.cell(r, 1, "Source: Master Benefits PRD Tab 5 — Activity Baseline for G&A Beneficial (GAB) · 72 FTEs (62 onshore, 10 offshore)").font = subtitle_font

    r = 3  # blank

    r = 4
    apply_section_row(ws, r, "GAB ACTIVITY BASELINE — % Time Allocation by Role", NUM_COLS)

    r = 5
    headers = ["Process", "AE-Existing (4)", "AM-Existing (27)", "AE-New (4)",
               "AM-New (10)", "Ben. Coordinator (5)", "Ben. UW (1)", "FTE Equiv.", "Initiative"]
    apply_header_row(ws, r, headers)

    activity_data = [
        ("CAP Build-out & Completion",      "10%", "10%", "20%", "20%", "10%", "—",   6.4, "A"),
        ("Benefits Booklet Preparation",     "5%",  "5%",  "5%",  "5%",  "15%", "—",   3.0, "A"),
        ("Benefits Renewal",                 "30%", "35%", "5%",  "5%",  "10%", "2%", 11.9, "A"),
        ("Sales Support & Prospecting",      "5%",  "5%",  "10%", "10%", "—",   "—",   3.0, "—"),
        ("Client Requirements Gathering",    "10%", "10%", "20%", "20%", "10%", "10%",  6.5, "—"),
        ("Benefits Underwriting & Quoting",  "5%",  "5%",  "10%", "10%", "30%", "75%",  5.2, "B"),
        ("Enrollment Education",             "10%", "10%", "10%", "10%", "—",   "5%",   4.6, "—"),
        ("Welcome Packet Build",             "—",   "—",   "5%",  "5%",  "10%", "—",    1.2, "—"),
        ("Broker Commission Verification",   "5%",  "5%",  "5%",  "5%",  "10%", "1%",   2.8, "—"),
        ("Strategic Client Outreach",        "15%", "10%", "5%",  "5%",  "—",   "5%",   4.1, "—"),
        ("QC & Work Audits",                 "5%",  "5%",  "5%",  "5%",  "5%",  "2%",   2.5, "—"),
    ]

    r = 6
    init_a_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    init_b_fill = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")

    for row_data in activity_data:
        fill = None
        if row_data[-1] == "A":
            fill = init_a_fill
        elif row_data[-1] == "B":
            fill = init_b_fill
        apply_data_row(ws, r, list(row_data), fill=fill)
        r += 1

    # ── Section 2: Automation Yield Summary ──
    r += 1
    apply_section_row(ws, r, "AUTOMATION YIELD PER INITIATIVE", NUM_COLS)
    r += 1

    yield_headers = ["Initiative", "Target Process", "Addressable FTEs", "Automation Yield",
                     "Est. FTE Recovery", "Key Metric", "Baseline", "Target", "Source"]
    apply_header_row(ws, r, yield_headers)
    r += 1

    yield_data = [
        ("A — CAP Migration", "CAP Build-out & Completion", "6.4 FTE equiv.", "40-50% of CAP/Renewal labor",
         "~3-4 FTEs", "Hours per CAP build", "2-4 hours", "<30 min (review only)", "PRD Tab 5"),
        ("A — CAP Migration", "Benefits Booklet Preparation", "3.0 FTE equiv.", "80-90% (auto-generated)",
         "~2.5 FTEs", "Booklet prep time", "Manual build", "Auto-generated from CAP", "PRD Tab 5"),
        ("A — CAP Migration", "Benefits Renewal", "11.9 FTE equiv.", "40-50% of renewal labor",
         "~5-6 FTEs", "Renewal CAP time", "Full manual rebuild", "Pre-filled, review only", "PRD Tab 5"),
        ("A — CAP Migration", "Ben Admin Setup (downstream)", "~4 FTEs (48% of Analyst time)", "20-30% rework eliminated",
         "~1-2 FTEs", "Setup rework cycles", "Frequent CAP errors", "Validated CAP data", "PRD Tab 4"),
        ("B — AI UW Assistant", "Benefits Underwriting & Quoting", "5.2 FTE equiv. (+ 20% UW capacity)", "60-70% of scrubbing time",
         "~3-4 FTEs", "Min per quote scrub", "20-30 min", "<5 min (review only)", "PRD Tab 4"),
        ("C — Carrier Portal", "Open-Market Enrollment", "~20% of Specialist time", "80-90% of entry time",
         "~2-3 FTEs", "Min per enrollment", "13-17 min", "<2 min (audit only)", "PRD Tab 4"),
        ("D — Service Automation", "WSE + Client Issue Resolution", "~35% of Specialist time", "30-40% deflection",
         "~4-5 FTEs", "ID card req/day", "2-6 (30-50 peak)", "Self-service", "PRD Tab 4"),
        ("E — Deduction Audit", "Payroll Reviews & Deductions", "~23% of Specialist time", "70-80% of review time",
         "~3-4 FTEs", "Min per client review", "20-60 min", "Exception review only", "PRD Tab 4"),
        ("F — Reconciliation", "Benefits Reconciliation", "5 FTEs (97% of Recon role)", "60-70% with stable Beacon",
         "~3-4 FTEs", "Platform reliability", "Partial (Excel fallback)", "Full production", "PRD Tab 4"),
    ]
    for row_data in yield_data:
        apply_data_row(ws, r, list(row_data))
        r += 1

    # ── Section 3: Total Impact Summary ──
    r += 1
    apply_section_row(ws, r, "TOTAL ADDRESSABLE IMPACT", NUM_COLS)
    r += 1

    impact_headers = ["Team", "Total FTEs", "Total FLC", "Onshore / Offshore",
                      "Directly Addressable FTEs", "Primary Initiatives", "", "", ""]
    apply_header_row(ws, r, impact_headers)
    r += 1

    impact_data = [
        ("G&A Beneficial (GAB)", "72 FTEs", "$8.2M", "62 onshore + 10 offshore",
         "31-32 FTEs", "A (CAP), B (AI UW)", "", "", ""),
        ("Benefits Administration", "98 FTEs", "$7.1M", "68 onshore + 30 offshore",
         "12-16 FTEs", "C (Portal), D (Service), E (Deduction), F (Recon)", "", "", ""),
    ]
    for row_data in impact_data:
        apply_data_row(ws, r, list(row_data))
        r += 1

    apply_data_row(ws, r, ["COMBINED TOTAL", "170 FTEs", "$15.3M", "130 onshore + 40 offshore",
                           "43-48 FTEs addressable", "All 6 Initiatives", "", "", ""],
                   font=bold_font, fill=total_fill)
    r += 1

    # Notes
    r += 1
    notes = [
        "Key assumptions:",
        "- Init A addresses 21.3 FTE equiv. in GAB (CAP Build-out 6.4 + Booklet 3.0 + Renewal 11.9) + ~10 offshore FTEs + downstream Ben Admin Analyst impact",
        "- Init B addresses 10,426 annual quotes at 20-30 min savings each = ~3,500-5,200 hours/year of scrubbing time recovered",
        "- Init C addresses 13-17 min per open-market enrollment across all non-EDI carrier enrollments",
        "- Initiatives D, E, F are vendor platform expansions with integration scope — impact projections from PRD Tab 4",
        "- FTE recovery does not mean headcount reduction — freed capacity is redirected to strategic client outreach, complex case handling, and business growth",
    ]
    for note in notes:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NUM_COLS)
        cell = ws.cell(r, 1, note)
        if note.startswith("Key"):
            cell.font = bold_font
        else:
            cell.font = note_font
        r += 1

    ws.freeze_panes = "A6"
    return ws


# =============================================================================
# MAIN
# =============================================================================
def main():
    wb = openpyxl.Workbook()

    build_migration_tracker(wb)
    build_sprint_plan(wb)
    build_effort_summary(wb)
    build_risk_register(wb)
    build_team_resources(wb)
    build_business_impact(wb)

    wb.save(OUTPUT_PATH)
    print(f"Saved to {OUTPUT_PATH}")
    print(f"Sheets: {wb.sheetnames}")

    # Quick validation
    wb2 = openpyxl.load_workbook(OUTPUT_PATH)
    for name in wb2.sheetnames:
        ws = wb2[name]
        print(f"  {name}: {ws.max_row} rows x {ws.max_column} cols")


if __name__ == "__main__":
    main()
